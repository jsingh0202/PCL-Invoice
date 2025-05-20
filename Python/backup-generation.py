from openpyxl import load_workbook, Workbook
from tkinter import Tk, filedialog
import os
import re
from copy import copy
from datetime import datetime

def get_date(backup):
    sheet = backup["WR1"]
    date_cell = sheet["L1"].value
    
    if date_cell:
        date = date_cell.strftime("%B %Y")
        return f"PCL Backup Export - {date}.xlsx"
    else:
        raise Exception("Date not found in cell L1.")
    
    
def save_output(export, backup):
    # save to output
    os.makedirs("out", exist_ok=True)
    output_file = get_date(backup)
    output_path = os.path.join("out", output_file)
    export.save(output_path)

    print("Saved output to: ", output_file)


def copy_styles(filtered_data, export_sheet, keep_cols, curr):
    for row in filtered_data:
        filtered_cells = [cell for i, cell in enumerate(row) if i in keep_cols]

        col_a = (
            str(row[0].value).strip().lower() if len(row) > 0 and row[0].value else ""
        )
        if "description" in col_a:
            continue

        if all(
            cell.value is None or str(cell.value).strip() == ""
            for cell in filtered_cells
        ):
            continue

        for col_i, cell in enumerate(filtered_cells, 1):
            new_cell = export_sheet.cell(row=curr, column=col_i, value=cell.value)
            if cell.has_style:
                new_cell.number_format = copy(cell.number_format)

        curr += 1
    
    return curr


def get_cols(filtered_data):
    # define pattern for cols to keep
    pattern = re.compile(
        r"Description|Total Contract Value|% Complete|Total Progress to Date|Previously Billed|Current Billing|Balance",
        re.IGNORECASE,
    )

    keep_cols = []
    for col_i in range(len(filtered_data[0])):
        col_v = [
            (
                str(row[col_i].value)
                if row[col_i] is not None and row[col_i].value is not None
                else ""
            )
            for row in filtered_data
        ]
        if any(pattern.search(val) for val in col_v):
            keep_cols.append(col_i)
    return keep_cols


def get_filtered(sheet):
    filtered_data = []
    for row in sheet.iter_rows(min_row=8, max_row=300, min_col=1, max_col=17):
        if all((cell.value is None or str(cell.value).strip() == "") for cell in row):
            continue

        col_a = (
            str(row[0].value).strip().lower() if len(row) > 0 and row[0].value else ""
        )
        col_b = (
            str(row[1].value).strip().lower() if len(row) > 1 and row[1].value else ""
        )
        if "work release #" in col_a or "services fee" in col_a or "profit and overhead" in col_a or "work release #" in col_b or "totals" in col_b:
            continue

        filtered_data.append(row)
    return filtered_data


def get_data(sheet):
    data = list(
        sheet.iter_rows(min_row=8, max_row=300, min_col=1, max_col=17, values_only=True)
    )
    return data


def create_export(backup, target_sheets, export_sheet):
    curr = 2
    for sheet in target_sheets:
        backup_sheet = backup[sheet]
        
        data = get_data(backup_sheet)
        if not data:
            continue

        filtered_data = get_filtered(backup_sheet)
        if not filtered_data:
            continue

        keep_cols = get_cols(filtered_data)
        curr = copy_styles(filtered_data, export_sheet, keep_cols, curr)


def add_headers(export_sheet):
    headers = [
        "Description",
        "Total Contract Value",
        "% Complete",
        "Total Progress to Date",
        "Previously Billed",
        "Current Billing",
        "Balance",
    ]

    for col_i, header in enumerate(headers, 1):
        export_sheet.cell(row=1, column=col_i, value=header)


def get_sheets(input):
    # Load workbook
    backup = load_workbook(input)
    sheets = backup.sheetnames

    # find sheets
    start = next((i for i, s in enumerate(sheets) if s.strip().lower() == "wr1"), None)
    end = next(
        (i for i, s in enumerate(sheets) if s.strip().lower() == "fixed fee"), None
    )

    if start is None or end is None:
        raise Exception("Start or end sheet not found.")
    if start > end:
        raise Exception("'FIXED FEE' appears before 'WR1'.")

    target_sheets = sheets[start : end + 1]
    print("Sheets to process: ", target_sheets)
    return backup, target_sheets


def get_file():
    # init filedialog
    Tk().withdraw()

    # File dialog
    print("Select your Excel file: ")
    input = filedialog.askopenfilename(
        title="Select Excel File", filetypes=[("Excel Files", "*xlsx")]
    )
    if not input:
        raise Exception("No file selected.")
    return input


def main():
    input = get_file()
    backup, sheets = get_sheets(input)

    # create new workbook
    export = Workbook()
    export_sheet = export.active
    export_sheet.title = "Export"

    add_headers(export_sheet)
    create_export(backup, sheets, export_sheet)
    save_output(export, backup)


if __name__ == "__main__":
    main()
