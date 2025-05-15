from openpyxl import load_workbook, Workbook
from tkinter import Tk, filedialog
import os
import re
from copy import copy
from datetime import datetime

# init filedialog
Tk().withdraw()

# File dialog
print("Select your Excel file: ")
input = filedialog.askopenfilename(
    title="Select Excel File", filetypes=[("Excel Files", "*xlsx")]
)
if not input:
    raise Exception("No file selected.")

# Load workbook
backup = load_workbook(input)
sheets = backup.sheetnames

# find sheets
start = next((i for i, s in enumerate(sheets) if s.strip().lower() == "wr1"), None)
end = next((i for i, s in enumerate(sheets) if s.strip().lower() == "fixed fee"), None)

if start is None or end is None:
    raise Exception("Start or end sheet not found.")
if start > end:
    raise Exception("'FIXED FEE' appears before 'WR1'.")

target_sheets = sheets[start : end + 1]
print("Sheets to process: ", target_sheets)

# define pattern for cols to keep
pattern = re.compile(
    r"Description|Total Contract Value|% Complete|Total Progress to Date|Previously Billed|Current Billing|Balance",
    re.IGNORECASE,
)

# create new workbook
export = Workbook()
export_sheet = export.active
export_sheet.title = "Export"

curr = 1

headers = [
    "Description",
    "Total Contract Value",
    "% Complete",
    "Total Progress to Date",
    "Previously Billed",
    "Current Billing",
    "Balance",
]

for col_i, header in enumerate(headers, 1):
    export_sheet.cell(row=curr, column=col_i, value=header)
curr += 1

for sheet in target_sheets:
    backup_sheet = backup[sheet]
    data = list(
        backup_sheet.iter_rows(
            min_row=8, max_row=300, min_col=1, max_col=17, values_only=True
        )
    )
    if not data:
        continue

    filtered_data = []
    for row in backup_sheet.iter_rows(min_row=8, max_row=300, min_col=1, max_col=17):
        if all((cell.value is None or str(cell.value).strip() == "") for cell in row):
            continue

        col_a = (
            str(row[0].value).strip().lower() if len(row) > 0 and row[0].value else ""
        )
        col_b = (
            str(row[1].value).strip().lower() if len(row) > 1 and row[1].value else ""
        )
        if "work release #" in col_a or "work release #" in col_b:
            continue

        filtered_data.append(row)

    keep_cols = []
    if filtered_data:
        for col_i in range(len(filtered_data[0])):
            col_v = [
                (
                    str(row[col_i].value)
                    if row[col_i] is not None and row[col_i].value is not None
                    else ""
                )
                for row in filtered_data
            ]
            if any(pattern.search(val) for val in col_v):
                keep_cols.append(col_i)

    for row in filtered_data:
        filtered_cells = [cell for i, cell in enumerate(row) if i in keep_cols]

        col_a = (
            str(row[0].value).strip().lower() if len(row) > 0 and row[0].value else ""
        )
        if "description" in col_a:
            continue

        if all(
            cell.value is None or str(cell.value).strip() == ""
            for cell in filtered_cells
        ):
            continue

        for col_i, cell in enumerate(filtered_cells, 1):
            new_cell = export_sheet.cell(row=curr, column=col_i, value=cell.value)
            if cell.has_style:
                new_cell.number_format = copy(cell.number_format)

        curr += 1

# save to output
os.makedirs("out", exist_ok=True)
timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
output_file = f"out/export_{timestamp}.xlsx"
export.save(output_file)

print("Saved output to: ", output_file)
