from openpyxl import load_workbook, Workbook
import os
import re
from copy import copy


def get_date(backup):
    """
    Gets the date from the backup sheet and formats it for the output file name.

    Args:
        backup (workbook): The backup workbook to get the date from.

    Raises:
        Exception: If the date is not found in cell L1.

    Returns:
        String: Formatted date string for the output file name.
    """
    sheet = backup["WR1"]
    date_cell = sheet["L1"].value

    if date_cell:
        date = date_cell.strftime("%B %Y")
        return f"PCL Backup Export - {date}.xlsx"
    else:
        raise Exception("Date not found in cell L1.")


def save_output(export, backup):
    """
    Generates the output file name and saves the export workbook.

    Args:
        export (workbook): The export workbook to save.
        backup (workbook): The backup workbook to get the date from.
    """
    # save to output
    os.makedirs("out", exist_ok=True)
    output_file = get_date(backup)
    output_path = os.path.join("out", output_file)
    export.save(output_path)

    print("Saved output to: ", output_file)


def get_cols(filtered_data):
    """
    Filters the data based on the columns to keep.

    Args:
        filtered_data (list): list of rows from the filtered data.

    Returns:
        list: list of data with the columns to keep.
    """
    # define pattern for cols to keep
    pattern = re.compile(
        r"Description|Total Contract Value|% Complete|Total Progress to Date|Previously Billed|Current Billing|Balance",
        re.IGNORECASE,
    )

    keep_cols = []
    for col_i in range(len(filtered_data[0])):
        col_v = [
            (
                str(row[col_i].value)
                if row[col_i] is not None and row[col_i].value is not None
                else ""
            )
            for row in filtered_data
        ]
        if any(pattern.search(val) for val in col_v):
            keep_cols.append(col_i)
    return keep_cols


def copy_styles(filtered_data, export_sheet, curr):
    """
    Copies the styles (specifically the number format) from the filtered data to the export sheet.
    Increments the current row index for each new row added to the export sheet.

    Args:
        filtered_data (list): list of rows from the filtered data.
        export_sheet (sheet): sheet to copy the styles to.
        curr (int): current row index in the export sheet.

    Returns:
        curr (int): updated current row index in the export sheet.
    """
    # Unpack just the rows for get_cols
    rows_only = [row for row, _ in filtered_data]
    keep_cols = get_cols(rows_only)
    
    for row, cell_location in filtered_data:
        filtered_cells = [cell for i, cell in enumerate(row) if i in keep_cols]

        col_a = (
            str(row[0].value).strip().lower() if len(row) > 0 and row[0].value else ""
        )
        if "description" in col_a:
            continue

        if all(
            cell.value is None or str(cell.value).strip() == ""
            for cell in filtered_cells
        ):
            continue
        
        # Write the index as the first column
        export_sheet.cell(row=curr, column=1, value=cell_location)
        
        for col_i, cell in enumerate(filtered_cells, 2):
            new_cell = export_sheet.cell(row=curr, column=col_i, value=cell.value)
            if cell.has_style:
                new_cell.number_format = copy(cell.number_format)

        curr += 1

    return curr


def get_filtered(sheet):
    """
    Filters the data from the current CoE backup sheet with constrains on the rows and columns.
    Also filters based on criteria on columns A and B.

    Args:
        sheet (sheet): The current CoE backup sheet to process.

    Returns:
        list: Filtered data from the sheet.
    """
    filtered_data = []
    blank_rows = 0
    last_blank = False
    for row in sheet.iter_rows(min_row=8, min_col=1, max_col=17):
        if blank_rows == 10:
            break
        if all((cell.value is None or str(cell.value).strip() == "") for cell in row):
            if last_blank:
                blank_rows += 1
            else:
                last_blank = True
            continue

        blank_rows = 0
        last_blank = False

        # Find the cell location of the first cell in the row
        cell_location = f"{sheet.title}-{row[0].coordinate}"
        filtered_data.append((row, cell_location))
    return filtered_data


def create_export(backup, target_sheet, export_sheet):
    """
    Creates the export sheet.

    Args:
        backup (workbook): The backup workbook.
        target_sheet (sheet): The sheet to process.
        export_sheet (sheet): The export sheet to write to.
    """
    curr = 2
    
    backup_sheet = backup[target_sheet]

    filtered_data = get_filtered(backup_sheet)
    if not filtered_data:
        raise Exception(f"No valid data found in sheet '{target_sheet}'.")

    curr = copy_styles(filtered_data, export_sheet, curr)


def add_headers(export_sheet):
    """
    Inserts headers into the export sheet.

    Args:
        export_sheet (workbook): The export sheet to add headers to.
    """
    headers = [
        "Index",
        "Description",
        "Total Contract Value",
        "% Complete",
        "Total Progress to Date",
        "Previously Billed",
        "Current Billing",
        "Balance",
    ]

    for col_i, header in enumerate(headers, 1):
        export_sheet.cell(row=1, column=col_i, value=header)


def get_sheet(input, sheet_name):
    """
    Takes in a file path and finds the sheet for processing.
    Returns the workbook and the sheet to process.

    Args:
        input (String): File path to the Excel file.

    Raises:
        Exception: If the sheet is not found.

    Returns:
        : Workbook and list of sheets to process.
    """
    # Load workbook
    backup = load_workbook(input)
    sheets = backup.sheetnames

    # find sheets
    sheet = next((i for i, s in enumerate(sheets) if s.strip().lower() == sheet_name), None)
    
    if sheet is None:
        raise Exception(f"Sheet '{sheet_name}' not found in the workbook.")
    

    target_sheets = sheets[sheet]
    print("Sheet to process: ", target_sheets)
    return backup, target_sheets


def generate_export(input):
    """
    Takes in a PCL CoE Backup file and produces a workable export file.
    """
    backup, sheets = get_sheet(input, "sov")

    # create new workbook
    export = Workbook()
    export_sheet = export.active
    export_sheet.title = "Export"

    add_headers(export_sheet)
    create_export(backup, sheets, export_sheet)

    return export
